import os
import numpy as np
import torch
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

def count_parameters(model, trainable=False):
    if trainable:
        return sum(p.numel() for p in model.parameters() if p.requires_grad)
    return sum(p.numel() for p in model.parameters())


def tensor2numpy(x):
    return x.cpu().data.numpy() if x.is_cuda else x.data.numpy()


def target2onehot(targets, n_classes):
    onehot = torch.zeros(targets.shape[0], n_classes).to(targets.device)
    onehot.scatter_(dim=1, index=targets.long().view(-1, 1), value=1.0)
    return onehot

def convert_time(seconds):
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    return "%d:%02d:%02d" % (h, m, s)

def makedirs(path):
    if not os.path.exists(path):
        os.makedirs(path)


def accuracy(y_pred, y_true, nb_old, increment):
    assert len(y_pred) == len(y_true), "Data length error."
    all_acc = {}
    all_acc["total"] = np.around(
        (y_pred == y_true).sum() * 100 / len(y_true), decimals=2
    )

    # Grouped accuracy
    for class_id in range(0, np.max(y_true), increment):
        idxes = np.where(
            np.logical_and(y_true >= class_id, y_true < class_id + increment)
        )[0]
        label = "{}-{}".format(
            str(class_id).rjust(2, "0"), str(class_id + increment - 1).rjust(2, "0")
        )
        all_acc[label] = np.around(
            (y_pred[idxes] == y_true[idxes]).sum() * 100 / len(idxes), decimals=2
        )

    # Old accuracy
    idxes = np.where(y_true < nb_old)[0]
    all_acc["old"] = (
        0 if len(idxes) == 0
        else np.around(
            (y_pred[idxes] == y_true[idxes]).sum() * 100 / len(idxes), decimals=2
        )
    )

    # New accuracy
    idxes = np.where(y_true >= nb_old)[0]
    all_acc["new"] = np.around(
        (y_pred[idxes] == y_true[idxes]).sum() * 100 / len(idxes), decimals=2
    )

    return all_acc


def split_images_labels(imgs):
    # split trainset.imgs in ImageFolder
    images = []
    labels = []
    for item in imgs:
        images.append(item[0])
        labels.append(item[1])

    return np.array(images), np.array(labels)

def list2dict(list):
    dict = {}
    for l in list:
        s = l.split(' ')
        id = int(s[0])
        cls = s[1]
        if id not in dict.keys():
            dict[id] = cls
        else:
            raise EOFError('The same ID can only appear once')
    return dict

def text_read(file):
    with open(file, 'r') as f:
        lines = f.readlines()
        for i, line in enumerate(lines):
            lines[i] = line.strip('\n')
    return lines




def save_results_to_excel(dataset_name, file_name, incremental_num, results, runing_time='', device='', note='', seed=''):
    base_dir = os.path.join(os.getcwd(), 'results', dataset_name, '{}'.format(str(incremental_num)))
    os.makedirs(base_dir, exist_ok=True)

    excel_path = os.path.join(base_dir, f"{file_name}.xlsx")

    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if os.path.exists(excel_path):
        book = load_workbook(excel_path)
    else:
        book = Workbook()
        book.remove(book.active)

    for metric, data in results.items():
        if metric not in book.sheetnames:
            sheet = book.create_sheet(metric)
            sheet.append(['Method', 'Timestamp', 'seed', 'Parameters'] + [f'task_{i}' for i in range(len(data[0][2]))] +['inc_acc','forget','grouped_top1_acc', 'running_time','device','note']) 
        else:
            sheet = book[metric]

        for entry in data:
            if entry != []:
                row = [entry[0], current_time, seed, entry[1]] + list(entry[2])   + [np.mean(entry[2]), str(entry[4]),str(entry[3]), runing_time, device, note]
                sheet.append(row)

    book.save(excel_path)

def get_device_name(device_type):
    device_names = []
    
    for device in device_type:
        if device == -1 or device == "-1":
            device_names.append("CPU")
        else:
            try:
                device_names.append(torch.cuda.get_device_name(int(device)))
            except Exception as e:
                pass
                # device_names.append(f"Error getting device name for cuda:{device}, {str(e)}")
    
    return device_names


if __name__ == "__main__":
    device_name = get_device_name([0,1])